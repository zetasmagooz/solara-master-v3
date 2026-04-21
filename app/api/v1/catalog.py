from io import BytesIO
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_current_user, get_db
from app.models.user import User
from app.schemas.catalog import (
    AttributeDefinitionCreate,
    AttributeDefinitionResponse,
    AttributeDefinitionUpdate,
    BrandCreate,
    BrandResponse,
    BrandUpdate,
    BulkImportRequest,
    BulkImportResponse,
    CategoryCreate,
    CategoryResponse,
    CategoryUpdate,
    CategoryWithSubcategories,
    PaginatedResponse,
    ProductAttributeCreate,
    ProductAttributeResponse,
    ProductCreate,
    ProductImageResponse,
    ProductImageUpdate,
    ProductImageUpload,
    ProductResponse,
    ProductUpdate,
    SubcategoryCreate,
    SubcategoryResponse,
    SubcategoryUpdate,
)
from app.services.ai_usage_service import consume_ai_usage, get_ai_image_cost, get_plan_features
from app.services.catalog_service import CatalogService
from app.services.image_gen_service import generate_kiosk_banner_image, generate_product_image

router = APIRouter(prefix="/catalog", tags=["catalog"])


# --- Categories ---
@router.get("/categories", response_model=list[CategoryWithSubcategories])
async def list_categories(
    store_id: Annotated[UUID, Query()],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Lista todas las categorías de una tienda con sus subcategorías incluidas.

    **Ejemplo curl:**
    ```bash
    curl -X GET "http://66.179.92.115:8005/api/v1/catalog/categories?store_id={store_id}" \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    return await service.get_categories(store_id, include_subcategories=True)


@router.post("/categories", response_model=CategoryResponse, status_code=status.HTTP_201_CREATED)
async def create_category(
    store_id: Annotated[UUID, Query()],
    data: CategoryCreate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Crea una nueva categoría para una tienda. Soporta imagen en base64.

    **Ejemplo curl:**
    ```bash
    curl -X POST "http://66.179.92.115:8005/api/v1/catalog/categories?store_id={store_id}" \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Bebidas", "description": "Refrescos y jugos"}'
    ```
    """
    service = CatalogService(db)
    payload = data.model_dump()
    # Handle base64 image
    if payload.get("image_url") and payload["image_url"].startswith("data:"):
        host_url = str(request.base_url).rstrip("/")
        try:
            payload["image_url"] = await service._save_image(payload["image_url"], "categories", host_url)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return await service.create_category(store_id, **payload)


@router.patch("/categories/{category_id}", response_model=CategoryResponse)
async def update_category(
    category_id: UUID,
    data: CategoryUpdate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Actualiza parcialmente una categoría. Soporta imagen en base64.

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/categories/{category_id} \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Bebidas Frías"}'
    ```
    """
    service = CatalogService(db)
    payload = data.model_dump(exclude_unset=True)
    if "image_url" in payload and payload["image_url"] and payload["image_url"].startswith("data:"):
        host_url = str(request.base_url).rstrip("/")
        try:
            payload["image_url"] = await service._save_image(payload["image_url"], "categories", host_url)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    category = await service.update_category(category_id, **payload)
    if not category:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Category not found")
    return category


@router.delete("/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_category(
    category_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Elimina una categoría por su ID. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X DELETE http://66.179.92.115:8005/api/v1/catalog/categories/{category_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    if not await service.delete_category(category_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Category not found")


# --- Subcategories ---
@router.post("/subcategories", response_model=SubcategoryResponse, status_code=status.HTTP_201_CREATED)
async def create_subcategory(
    store_id: Annotated[UUID, Query()],
    data: SubcategoryCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Crea una nueva subcategoría asociada a una categoría existente.

    **Ejemplo curl:**
    ```bash
    curl -X POST "http://66.179.92.115:8005/api/v1/catalog/subcategories?store_id={store_id}" \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Refrescos", "category_id": "{category_id}"}'
    ```
    """
    service = CatalogService(db)
    return await service.create_subcategory(store_id, **data.model_dump())


@router.get("/categories/{category_id}/subcategories", response_model=list[SubcategoryResponse])
async def list_subcategories(
    category_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Lista las subcategorías de una categoría específica.

    **Ejemplo curl:**
    ```bash
    curl -X GET http://66.179.92.115:8005/api/v1/catalog/categories/{category_id}/subcategories \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    return await service.get_subcategories(category_id)


@router.patch("/subcategories/{subcategory_id}", response_model=SubcategoryResponse)
async def update_subcategory(
    subcategory_id: UUID,
    data: SubcategoryUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Actualiza parcialmente una subcategoría. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/subcategories/{subcategory_id} \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Refrescos de Cola"}'
    ```
    """
    service = CatalogService(db)
    sub = await service.update_subcategory(subcategory_id, **data.model_dump(exclude_unset=True))
    if not sub:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Subcategory not found")
    return sub


@router.delete("/subcategories/{subcategory_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_subcategory(
    subcategory_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Elimina una subcategoría por su ID. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X DELETE http://66.179.92.115:8005/api/v1/catalog/subcategories/{subcategory_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    if not await service.delete_subcategory(subcategory_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Subcategory not found")


# --- Brands ---
@router.get("/brands", response_model=list[BrandResponse])
async def list_brands(
    store_id: Annotated[UUID, Query()],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Lista todas las marcas de una tienda.

    **Ejemplo curl:**
    ```bash
    curl -X GET "http://66.179.92.115:8005/api/v1/catalog/brands?store_id={store_id}" \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    return await service.get_brands(store_id)


@router.post("/brands", response_model=BrandResponse, status_code=status.HTTP_201_CREATED)
async def create_brand(
    store_id: Annotated[UUID, Query()],
    data: BrandCreate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Crea una nueva marca para una tienda. Soporta imagen en base64.

    **Ejemplo curl:**
    ```bash
    curl -X POST "http://66.179.92.115:8005/api/v1/catalog/brands?store_id={store_id}" \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Coca Cola"}'
    ```
    """
    service = CatalogService(db)
    payload = data.model_dump()
    # Handle base64 image if provided
    base64_image = payload.pop("image_url", None)
    brand = await service.create_brand(store_id, **payload)
    if base64_image and base64_image.startswith("data:"):
        host_url = str(request.base_url).rstrip("/")
        try:
            url = await service._save_image(base64_image, "brands", host_url)
            brand = await service.update_brand(brand.id, image_url=url)
        except ValueError:
            pass
    return brand


@router.get("/brands/{brand_id}", response_model=BrandResponse)
async def get_brand(
    brand_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Obtiene una marca por su ID. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X GET http://66.179.92.115:8005/api/v1/catalog/brands/{brand_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    brand = await service.get_brand(brand_id)
    if not brand:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Brand not found")
    return brand


@router.patch("/brands/{brand_id}", response_model=BrandResponse)
async def update_brand(
    brand_id: UUID,
    data: BrandUpdate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Actualiza parcialmente una marca. Soporta imagen en base64.

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/brands/{brand_id} \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Coca-Cola Company"}'
    ```
    """
    service = CatalogService(db)
    payload = data.model_dump(exclude_unset=True)
    # Handle base64 image if provided
    if "image_url" in payload and payload["image_url"] and payload["image_url"].startswith("data:"):
        host_url = str(request.base_url).rstrip("/")
        try:
            payload["image_url"] = await service._save_image(payload["image_url"], "brands", host_url)
        except ValueError as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    brand = await service.update_brand(brand_id, **payload)
    if not brand:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Brand not found")
    return brand


@router.delete("/brands/{brand_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_brand(
    brand_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Elimina una marca por su ID. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X DELETE http://66.179.92.115:8005/api/v1/catalog/brands/{brand_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    if not await service.delete_brand(brand_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Brand not found")


# --- Products ---
@router.get("/products/trending")
async def trending_products(
    store_id: Annotated[UUID, Query()],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    limit: int = Query(10, ge=1, le=50),
):
    """Retorna los IDs de los productos mas vendidos de una tienda, ordenados por popularidad.

    **Ejemplo curl:**
    ```bash
    curl -X GET "http://66.179.92.115:8005/api/v1/catalog/products/trending?store_id={store_id}&limit=10" \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    return await service.get_trending_ids(store_id, limit=limit)


@router.get("/products", response_model=PaginatedResponse[ProductResponse])
async def list_products(
    store_id: Annotated[UUID, Query()],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: str | None = None,
    category_id: UUID | None = None,
    brand_id: UUID | None = None,
    is_active: bool | None = None,
    low_stock: bool = False,
    is_favorite: bool | None = None,
    subcategory_id: UUID | None = None,
):
    """Lista productos paginados con filtros opcionales (busqueda, categoria, marca, stock bajo, favoritos).

    **Ejemplo curl:**
    ```bash
    curl -X GET "http://66.179.92.115:8005/api/v1/catalog/products?store_id={store_id}&page=1&per_page=20&search=coca" \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    return await service.get_products_paginated(
        store_id,
        page=page,
        per_page=per_page,
        search=search,
        category_id=category_id,
        brand_id=brand_id,
        is_active=is_active,
        low_stock=low_stock,
        is_favorite=is_favorite,
        subcategory_id=subcategory_id,
    )


@router.post("/products", response_model=ProductResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    store_id: Annotated[UUID, Query()],
    data: ProductCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Crea un nuevo producto para una tienda. Soporta atributos opcionales (talla, color, etc.).

    **Ejemplo curl:**
    ```bash
    curl -X POST "http://66.179.92.115:8005/api/v1/catalog/products?store_id={store_id}" \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{
        "name": "Coca Cola 600ml",
        "price": 25.00,
        "cost": 18.50,
        "stock": 50,
        "category_id": "{category_id}"
      }'
    ```
    """
    service = CatalogService(db)
    payload = data.model_dump()
    attributes_data = [a.model_dump() for a in data.attributes] if data.attributes else []
    payload.pop("attributes", None)

    if attributes_data:
        return await service.create_product_with_attributes(store_id, attributes_data, **payload)
    return await service.create_product(store_id, **payload)


# --- Bulk Import ---
@router.get("/products/import-template")
async def download_import_template():
    """Genera y descarga una plantilla Excel para importacion masiva de productos.

    **Ejemplo curl:**
    ```bash
    curl -X GET http://66.179.92.115:8005/api/v1/catalog/products/import-template \\
      -o plantilla_productos.xlsx
    ```
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Sheet 1: Productos ---
    ws = wb.active
    ws.title = "Productos"

    headers = [
        ("Nombre*", True),
        ("Precio de Venta*", True),
        ("Descripción", False),
        ("SKU", False),
        ("Código de Barras", False),
        ("Precio de Costo", False),
        ("Stock", False),
        ("Stock Mínimo", False),
        ("Stock Máximo", False),
        ("Categoría", False),
        ("Subcategoría", False),
        ("Marca", False),
        ("Fecha de Caducidad", False),
        ("Mostrar en POS", False),
        ("Mostrar en Kiosko", False),
    ]

    orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    blue_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [25, 18, 35, 15, 18, 18, 12, 14, 14, 20, 20, 20, 18, 16, 16]

    for col_idx, (header_text, is_required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = orange_fill if is_required else blue_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]

    # Example row
    example = [
        "Coca Cola 600ml", 25.00, "Refresco de cola 600ml", "COC-600",
        "7501055300846", 18.50, 50, 10, 100, "Bebidas", "Refrescos",
        "Coca Cola", "2026-12-31", "SI", "SI",
    ]
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=val)

    ws.row_dimensions[1].height = 30

    # --- Sheet 2: Instrucciones ---
    ws2 = wb.create_sheet("Instrucciones")
    ws2.column_dimensions["A"].width = 80

    instructions = [
        "INSTRUCCIONES PARA LLENAR LA PLANTILLA",
        "",
        "Campos obligatorios (marcados con *) — columnas en naranja:",
        "  • Nombre*: Nombre del producto (texto)",
        "  • Precio de Venta*: Precio al público (número mayor a 0)",
        "",
        "Campos opcionales — columnas en azul:",
        "  • Descripción: Descripción del producto",
        "  • SKU: Código interno del producto",
        "  • Código de Barras: Código de barras del producto",
        "  • Precio de Costo: Precio de compra (número)",
        "  • Stock: Cantidad en inventario (número, default: 0)",
        "  • Stock Mínimo: Cantidad mínima de alerta (número, default: 0)",
        "  • Stock Máximo: Cantidad máxima permitida (número)",
        "  • Categoría: Nombre de la categoría (se crea automáticamente si no existe)",
        "  • Subcategoría: Nombre de la subcategoría (requiere Categoría, se crea automáticamente)",
        "  • Marca: Nombre de la marca (se crea automáticamente si no existe)",
        "  • Fecha de Caducidad: Formato YYYY-MM-DD (ejemplo: 2026-12-31)",
        "  • Mostrar en POS: SI o NO (default: SI)",
        "  • Mostrar en Kiosko: SI o NO (default: SI)",
        "",
        "NOTAS IMPORTANTES:",
        "  • La fila 1 contiene los encabezados — NO la modifiques",
        "  • La fila 2 tiene un ejemplo — puedes borrarla o reemplazarla",
        "  • Empieza a llenar desde la fila 2 (o fila 3 si dejas el ejemplo)",
        "  • Límite máximo: 10,000 productos por importación",
        "  • Las imágenes se agregan después desde la app (individual o con IA)",
        "  • Las categorías, subcategorías y marcas se crean automáticamente si no existen",
    ]
    title_font = Font(bold=True, size=14)
    normal_font = Font(size=11)
    for i, line in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=line)
        cell.font = title_font if i == 1 else normal_font

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"},
    )


@router.get("/products/export-template")
async def export_products_template(
    store_id: Annotated[UUID, Query()],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Genera plantilla Excel pre-llenada con los productos existentes de la tienda."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    service = CatalogService(db)
    products = await service.get_all_products_for_export(store_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        ("ID", False),
        ("Nombre*", True),
        ("Precio de Venta*", True),
        ("Descripción", False),
        ("SKU", False),
        ("Código de Barras", False),
        ("Precio de Costo", False),
        ("Stock", False),
        ("Stock Mínimo", False),
        ("Stock Máximo", False),
        ("Categoría", False),
        ("Subcategoría", False),
        ("Marca", False),
        ("Fecha de Caducidad", False),
        ("Mostrar en POS", False),
        ("Mostrar en Kiosko", False),
    ]

    orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    blue_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [36, 25, 18, 35, 15, 18, 18, 12, 14, 14, 20, 20, 20, 18, 16, 16]

    for col_idx, (header_text, is_required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        if header_text == "ID":
            cell.fill = gray_fill
        elif is_required:
            cell.fill = orange_fill
        else:
            cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]

    ws.row_dimensions[1].height = 30

    # Fill product rows
    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=str(p["id"]))
        ws.cell(row=row_idx, column=2, value=p["name"])
        ws.cell(row=row_idx, column=3, value=p["base_price"])
        ws.cell(row=row_idx, column=4, value=p.get("description") or "")
        ws.cell(row=row_idx, column=5, value=p.get("sku") or "")
        ws.cell(row=row_idx, column=6, value=p.get("barcode") or "")
        ws.cell(row=row_idx, column=7, value=p.get("cost_price") or "")
        ws.cell(row=row_idx, column=8, value=p.get("stock", 0))
        ws.cell(row=row_idx, column=9, value=p.get("min_stock", 0))
        ws.cell(row=row_idx, column=10, value=p.get("max_stock") or "")
        ws.cell(row=row_idx, column=11, value=p.get("category_name") or "")
        ws.cell(row=row_idx, column=12, value=p.get("subcategory_name") or "")
        ws.cell(row=row_idx, column=13, value=p.get("brand_name") or "")
        ws.cell(row=row_idx, column=14, value=str(p["expiry_date"]) if p.get("expiry_date") else "")
        ws.cell(row=row_idx, column=15, value="SI" if p.get("show_in_pos", True) else "NO")
        ws.cell(row=row_idx, column=16, value="SI" if p.get("show_in_kiosk", True) else "NO")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos_tienda.xlsx"},
    )


@router.post("/products/bulk-import", response_model=BulkImportResponse)
async def bulk_import_products(
    store_id: Annotated[UUID, Query()],
    data: BulkImportRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Importa productos masivamente desde datos parseados de Excel. Opcionalmente genera imagenes con IA.

    **Ejemplo curl:**
    ```bash
    curl -X POST "http://66.179.92.115:8005/api/v1/catalog/products/bulk-import?store_id={store_id}" \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{
        "products": [
          {"name": "Coca Cola 600ml", "price": 25.00, "stock": 50},
          {"name": "Pepsi 600ml", "price": 22.00, "stock": 30}
        ],
        "generate_images": false
      }'
    ```
    """
    service = CatalogService(db)
    host_url = str(request.base_url).rstrip("/")
    rows = [r.model_dump() for r in data.products]
    result = await service.bulk_import_products(store_id, rows, host_url, data.generate_images)

    # Generate images in background if requested
    if data.generate_images and result["created_product_ids"]:
        async def _gen_images(product_ids: list[str], base_url: str):
            from app.database import AsyncSessionLocal
            import base64 as b64mod
            async with AsyncSessionLocal() as bg_db:
                bg_service = CatalogService(bg_db)
                for pid in product_ids:
                    try:
                        product = await bg_service.get_product(UUID(pid))
                        if product:
                            jpeg_bytes = await generate_product_image(product.name, product.description)
                            b64_data = f"data:image/jpeg;base64,{b64mod.b64encode(jpeg_bytes).decode()}"
                            await bg_service.save_product_image(UUID(pid), b64_data, True, base_url)
                    except Exception:
                        pass
                await bg_db.commit()

        background_tasks.add_task(_gen_images, result["created_product_ids"], host_url)

    return result


@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Obtiene un producto por su ID con imagenes y atributos. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X GET http://66.179.92.115:8005/api/v1/catalog/products/{product_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    product = await service.get_product(product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    return product


@router.patch("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: UUID,
    data: ProductUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Actualiza parcialmente un producto (nombre, precio, stock, etc.). Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/products/{product_id} \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"price": 30.00, "stock": 100}'
    ```
    """
    service = CatalogService(db)
    product = await service.update_product(product_id, **data.model_dump(exclude_unset=True))
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    return product


@router.delete("/products/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_product(
    product_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Elimina un producto por su ID. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X DELETE http://66.179.92.115:8005/api/v1/catalog/products/{product_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    if not await service.delete_product(product_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")


@router.patch("/products/{product_id}/favorite", response_model=ProductResponse)
async def toggle_product_favorite(
    product_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Alterna el estado de favorito de un producto. Retorna el producto actualizado.

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/products/{product_id}/favorite \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    try:
        return await service.toggle_favorite(product_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")


@router.put("/products/{product_id}/attributes", response_model=list[ProductAttributeResponse])
async def upsert_product_attributes(
    product_id: UUID,
    data: list[ProductAttributeCreate],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Crea o reemplaza los atributos de un producto (talla, color, etc.).

    **Ejemplo curl:**
    ```bash
    curl -X PUT http://66.179.92.115:8005/api/v1/catalog/products/{product_id}/attributes \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '[{"definition_id": "{attr_def_id}", "value": "Grande"}]'
    ```
    """
    service = CatalogService(db)
    attrs = [a.model_dump() for a in data]
    return await service.set_product_attributes(product_id, attrs)


# --- Product Images ---
@router.post("/products/{product_id}/images", response_model=ProductImageResponse, status_code=status.HTTP_201_CREATED)
async def upload_product_image(
    product_id: UUID,
    data: ProductImageUpload,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Sube una imagen en base64 para un producto. Puede marcarse como imagen principal.

    **Ejemplo curl:**
    ```bash
    curl -X POST http://66.179.92.115:8005/api/v1/catalog/products/{product_id}/images \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"base64_data": "data:image/jpeg;base64,/9j/4AAQ...", "is_primary": true}'
    ```
    """
    service = CatalogService(db)
    product = await service.get_product(product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    host_url = str(request.base_url).rstrip("/")
    try:
        return await service.save_product_image(product_id, data.base64_data, data.is_primary, host_url)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/products/{product_id}/generate-image", response_model=ProductImageResponse, status_code=status.HTTP_201_CREATED)
async def generate_product_image_endpoint(
    product_id: UUID,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    """Genera una imagen con IA para un producto a partir de su nombre y descripcion.
    Cobra usos del contador diario de IA (configurable por plan, default 5).

    **Ejemplo curl:**
    ```bash
    curl -X POST http://66.179.92.115:8005/api/v1/catalog/products/{product_id}/generate-image \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    product = await service.get_product(product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")

    features = await get_plan_features(db, current_user.organization_id)
    cost = get_ai_image_cost(features)
    await consume_ai_usage(db, current_user.organization_id, cost=cost)

    try:
        jpeg_bytes = await generate_product_image(product.name, product.description)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Image generation failed: {e}")

    import base64 as b64mod
    base64_data = f"data:image/jpeg;base64,{b64mod.b64encode(jpeg_bytes).decode()}"
    host_url = str(request.base_url).rstrip("/")
    return await service.save_product_image(product_id, base64_data, True, host_url)


class CatalogImageGenerateRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    description: str | None = Field(default=None, max_length=500)
    orientation: str = Field(default="square", pattern="^(square|portrait|wide_banner)$")


class CatalogImageGenerateResponse(BaseModel):
    image_url: str  # data:image/jpeg;base64,... — el frontend lo envía en el payload del create/update
    ai_cost: int
    ai_used: int
    ai_limit: int


@router.post("/ai/generate-image", response_model=CatalogImageGenerateResponse)
async def generate_catalog_image(
    data: CatalogImageGenerateRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    """Genera una imagen con IA para categoría/marca a partir de nombre y descripción.
    No persiste — retorna base64 data URL que el frontend incluye en el payload de create/update.
    Cobra usos del contador diario de IA (configurable por plan, default 5).

    **Ejemplo curl:**
    ```bash
    curl -X POST http://66.179.92.115:8005/api/v1/catalog/ai/generate-image \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Bebidas frías", "description": "Refrescos y jugos"}'
    ```
    """
    features = await get_plan_features(db, current_user.organization_id)
    cost = get_ai_image_cost(features)
    used, limit = await consume_ai_usage(db, current_user.organization_id, cost=cost)

    try:
        jpeg_bytes = await generate_kiosk_banner_image(
            data.name, data.description, orientation=data.orientation
        )
    except Exception as e:
        # get_db hace rollback al propagar la excepción → no se cobran los usos
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Image generation failed: {e}")

    import base64 as b64mod
    base64_data = f"data:image/jpeg;base64,{b64mod.b64encode(jpeg_bytes).decode()}"
    return CatalogImageGenerateResponse(
        image_url=base64_data,
        ai_cost=cost,
        ai_used=used,
        ai_limit=limit,
    )


@router.delete("/product-images/{image_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_product_image(
    image_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Elimina una imagen de producto por su ID. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X DELETE http://66.179.92.115:8005/api/v1/catalog/product-images/{image_id} \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    if not await service.delete_product_image(image_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Image not found")


@router.patch("/product-images/{image_id}", response_model=ProductImageResponse)
async def update_product_image(
    image_id: UUID,
    data: ProductImageUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Actualiza una imagen de producto (ej. establecerla como principal).

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/product-images/{image_id} \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"is_primary": true}'
    ```
    """
    service = CatalogService(db)
    if data.is_primary:
        image = await service.set_primary_image(image_id)
    else:
        image = None
    if not image:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Image not found")
    return image


# --- Attribute Definitions ---
@router.get("/attribute-definitions", response_model=list[AttributeDefinitionResponse])
async def list_attribute_definitions(
    store_id: Annotated[UUID, Query()],
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Lista las definiciones de atributos de una tienda (talla, color, material, etc.).

    **Ejemplo curl:**
    ```bash
    curl -X GET "http://66.179.92.115:8005/api/v1/catalog/attribute-definitions?store_id={store_id}" \\
      -H "Authorization: Bearer {token}"
    ```
    """
    service = CatalogService(db)
    return await service.get_attribute_definitions(store_id)


@router.post("/attribute-definitions", response_model=AttributeDefinitionResponse, status_code=status.HTTP_201_CREATED)
async def create_attribute_definition(
    store_id: Annotated[UUID, Query()],
    data: AttributeDefinitionCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Crea una nueva definicion de atributo para una tienda (nombre, tipo, opciones).

    **Ejemplo curl:**
    ```bash
    curl -X POST "http://66.179.92.115:8005/api/v1/catalog/attribute-definitions?store_id={store_id}" \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Talla", "type": "select", "options": ["S", "M", "L", "XL"]}'
    ```
    """
    service = CatalogService(db)
    return await service.create_attribute_definition(store_id, **data.model_dump())


@router.patch("/attribute-definitions/{definition_id}", response_model=AttributeDefinitionResponse)
async def update_attribute_definition(
    definition_id: UUID,
    data: AttributeDefinitionUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
):
    """Actualiza una definicion de atributo existente. Retorna 404 si no existe.

    **Ejemplo curl:**
    ```bash
    curl -X PATCH http://66.179.92.115:8005/api/v1/catalog/attribute-definitions/{definition_id} \\
      -H "Authorization: Bearer {token}" \\
      -H "Content-Type: application/json" \\
      -d '{"name": "Talla Ropa", "options": ["XS", "S", "M", "L", "XL"]}'
    ```
    """
    service = CatalogService(db)
    ad = await service.update_attribute_definition(definition_id, **data.model_dump(exclude_unset=True))
    if not ad:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attribute definition not found")
    return ad
